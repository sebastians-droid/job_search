import os
import re
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import sys

# lxml is 3-5x faster than html.parser for BeautifulSoup; fall back gracefully
try:
    import lxml  # noqa: F401
    _HTML_PARSER = 'lxml'
except ImportError:
    _HTML_PARSER = 'html.parser'

# ── CONFIGURATION ──────────────────────────────────────────────────────────────
LOGIN_URL            = "https://ui.bidx.com/login"
LOGIN_FILE           = "bidx_login.txt"
MILLING_CONFIG       = "milling_config.xlsx"
GRINDING_CONFIG      = "grinding_grooving_config.xlsx"
MILLING_OUTPUT       = "bidx_results_milling.xlsx"
GRINDING_OUTPUT      = "bidx_results_grinding_grooving.xlsx"
RESTART_EVERY_N_DOTS = 8
HEADLESS             = os.environ.get('BIDX_HEADLESS', 'true').lower() not in ('0', 'false', 'no')

# Column definitions for each output type
_MILLING_COLS  = ['letting_date', 'proposal_id', 'district', 'project_description',
                   'item_number', 'description', 'unit', 'quantity']
_MILLING_HDRS  = ["Letting Date", "Proposal ID", "District", "Project Description",
                   "Item Number", "Description", "Unit", "Quantity"]
_GRINDING_COLS = ['type'] + _MILLING_COLS
_GRINDING_HDRS = ["Type"] + _MILLING_HDRS


# ── CREDENTIALS ────────────────────────────────────────────────────────────────

def load_credentials():
    """Load credentials from env vars (Azure Key Vault) or local login file."""
    username = os.environ.get('BIDX_USERNAME', '').strip()
    password = os.environ.get('BIDX_PASSWORD', '').strip()
    if username and password:
        return username, password

    if not os.path.exists(LOGIN_FILE):
        print("ERROR: BIDX credentials not found.")
        print("  Set BIDX_USERNAME and BIDX_PASSWORD environment variables,")
        print(f"  or create '{LOGIN_FILE}' (format: username|password) for local dev.")
        sys.exit(1)
    try:
        with open(LOGIN_FILE) as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                parts = line.split('|')
                if len(parts) == 2:
                    return parts[0].strip(), parts[1].strip()
        print(f"ERROR: No valid credentials found in '{LOGIN_FILE}'.")
        sys.exit(1)
    except Exception as e:
        print(f"ERROR reading login file: {e}")
        sys.exit(1)


# ── CONFIG LOADING ─────────────────────────────────────────────────────────────

def _load_excel_config(path, required_cols):
    """Generic Excel config loader — returns a DataFrame or exits."""
    if not os.path.exists(path):
        print(f"ERROR: Config file '{path}' not found.")
        sys.exit(1)
    try:
        df = pd.read_excel(path, sheet_name='Config', dtype=str)
        df.columns = [c.strip() for c in df.columns]
        missing = required_cols - set(df.columns)
        if missing:
            print(f"ERROR: '{path}' is missing columns: {missing}")
            sys.exit(1)
        return df
    except SystemExit:
        raise
    except Exception as e:
        print(f"ERROR reading '{path}': {e}")
        sys.exit(1)


def load_milling_config():
    """Return dict keyed by url_slug with milling keyword lists (uppercase)."""
    df = _load_excel_config(MILLING_CONFIG,
                            {'ID', 'DOT Name', 'DOT URL Slug', 'Milling Keywords'})
    result = {}
    for _, row in df.iterrows():
        slug    = str(row['DOT URL Slug']).strip()
        raw_kw  = str(row['Milling Keywords']).strip()
        keywords = (
            [] if raw_kw.upper() in ('NONE', 'NAN', '')
            else [k.strip().upper() for k in raw_kw.split(',') if k.strip()]
        )
        result[slug] = {
            'name':     str(row['DOT Name']).strip(),
            'url_slug': slug,
            'keywords': keywords,
        }
    return result


def load_grinding_config():
    """Return dict keyed by url_slug with grinding/grooving item number lists."""
    df = _load_excel_config(GRINDING_CONFIG,
                            {'ID', 'DOT Name', 'DOT URL Slug', 'Grinding Items', 'Grooving Items'})
    result = {}
    for _, row in df.iterrows():
        slug = str(row['DOT URL Slug']).strip()

        def parse_items(raw):
            raw = str(raw).strip()
            return ([] if raw.upper() in ('NONE', 'NAN', '')
                    else [x.strip() for x in raw.split(',') if x.strip()])

        result[slug] = {
            'name':           str(row['DOT Name']).strip(),
            'url_slug':       slug,
            'grinding_items': parse_items(row['Grinding Items']),
            'grooving_items': parse_items(row['Grooving Items']),
        }
    return result


def build_dot_list(milling_dots, grinding_dots):
    """Merge both config dicts into a single ordered list of DOTs to scrape."""
    # Use milling order as primary; append any grinding-only DOTs at the end
    seen = {}
    for slug, m in milling_dots.items():
        seen[slug] = m['name']
    for slug, g in grinding_dots.items():
        if slug not in seen:
            seen[slug] = g['name']

    dots = []
    for slug, name in seen.items():
        dots.append({
            'name':            name,
            'url_slug':        slug,
            'milling_keywords': milling_dots[slug]['keywords']        if slug in milling_dots  else [],
            'grinding_items':   grinding_dots[slug]['grinding_items'] if slug in grinding_dots else [],
            'grooving_items':   grinding_dots[slug]['grooving_items'] if slug in grinding_dots else [],
        })
    return dots


# ── BROWSER / LOGIN ────────────────────────────────────────────────────────────

def login_to_bidx(driver, username, password, retry_count=0):
    """Automated BIDX login with one retry on failure."""
    try:
        driver.get(LOGIN_URL)
        wait_for_page_load(driver, timeout=15)
        time.sleep(3)

        email_input = WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[aria-label="Email"]'))
        )
        email_input.clear()
        time.sleep(0.5)
        email_input.send_keys(username)

        password_input = driver.find_element(By.CSS_SELECTOR, 'input[aria-label="Password"]')
        password_input.clear()
        time.sleep(0.5)
        password_input.send_keys(password)

        driver.find_element(By.CSS_SELECTOR, 'cw-button[type="submit"]').click()
        time.sleep(5)
        wait_for_page_load(driver, timeout=20)

        if '/login' not in driver.current_url:
            return True

        print(f"Warning: Login may have failed. URL: {driver.current_url}")
        if retry_count < 1:
            print("Retrying login in 5 seconds...")
            time.sleep(5)
            return login_to_bidx(driver, username, password, retry_count + 1)
        return False
    except Exception as e:
        print(f"Login error: {e}")
        if retry_count < 1:
            time.sleep(5)
            return login_to_bidx(driver, username, password, retry_count + 1)
        return False


def initialize_browser(username, password, is_restart=False):
    """Open Chrome, perform automated login, and return the driver."""
    label = "Restarting browser" if is_restart else "Initializing Chrome browser"
    print(f"{label}...")

    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--disable-extensions')
    options.add_argument('--no-first-run')
    options.add_argument('--no-default-browser-check')
    # Images are irrelevant for DOM scraping — blocking them speeds up every page load
    options.add_argument('--blink-settings=imagesEnabled=false')
    if HEADLESS:
        options.add_argument('--headless=new')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')

    chrome_bin = os.environ.get('CHROME_BIN', '').strip()
    if chrome_bin:
        options.binary_location = chrome_bin

    driver = webdriver.Chrome(options=options)

    if is_restart:
        print("Waiting 5 seconds before login (avoid rate limiting)...")
        time.sleep(5)

    if not login_to_bidx(driver, username, password):
        print("ERROR: Automated login failed after retries!")
        driver.quit()
        sys.exit(1)

    print(f"OK - {'Restarted and logged in' if is_restart else 'Logged in successfully'}\n")
    return driver


# ── PAGE UTILITIES ─────────────────────────────────────────────────────────────

def wait_for_page_load(driver, timeout=30):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script('return document.readyState') == 'complete'
        )
        return True
    except:
        return False


def wait_for_elements_stable(driver, css_selector, timeout=30, check_interval=0.5):
    """Block until element count stops changing, indicating the page is done loading."""
    try:
        start_time     = time.time()
        previous_count = 0
        stable_count   = 0
        while time.time() - start_time < timeout:
            current_count = len(driver.find_elements(By.CSS_SELECTOR, css_selector))
            if current_count > 0 and current_count == previous_count:
                stable_count += 1
                if stable_count >= 2:
                    return True
            else:
                stable_count = 0
            previous_count = current_count
            time.sleep(check_interval)
        return previous_count > 0
    except:
        return False


# ── SCRAPING ───────────────────────────────────────────────────────────────────

def extract_lettings_from_page(driver):
    """Return list of {letting_date, letting_id, district} from the upcoming-lettings table."""
    print("Extracting lettings from table...")
    try:
        WebDriverWait(driver, 15).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "a[href*='/lettings/']"))
        )
    except:
        print("OK - Found 0 lettings\n")
        return []

    # Use .+ so letting IDs with hyphens/dots are captured correctly
    js_code = """
    const upcomingCard = document.querySelector('cw-card[data-cy="upcoming-table"]');
    if (!upcomingCard) return [];
    const links = upcomingCard.querySelectorAll("a[href*='/lettings/']");
    const lettings = [];
    const seenIds = new Set();
    links.forEach(link => {
        const href = link.getAttribute('href');
        if (!href) return;
        const match = href.match(/\\/lettings\\/(.+)$/);
        if (!match) return;
        const lettingId = match[1];
        if (seenIds.has(lettingId)) return;
        seenIds.add(lettingId);
        const dateText = link.textContent.trim();
        let district = '';
        const parentSpan = link.closest('span[slot^="date-"]');
        if (parentSpan) {
            const allSpans = parentSpan.parentElement.querySelectorAll('span');
            let foundDate = false;
            for (let span of allSpans) {
                if (foundDate && span.textContent.length < 10) { district = span.textContent.trim(); break; }
                if (span === parentSpan) foundDate = true;
            }
        }
        if (dateText && lettingId) lettings.push({ date: dateText, lettingId: lettingId, district: district });
    });
    return lettings;
    """
    lettings_data = driver.execute_script(js_code)
    lettings = [{'letting_date': x['date'], 'letting_id': x['lettingId'], 'district': x['district']}
                for x in lettings_data]
    print(f"OK - Found {len(lettings)} lettings\n")
    return lettings


def navigate_to_letting_and_extract_proposals(driver, base_url, letting_id):
    """Return list of {proposal_id, district, description, detail_url} for a letting."""
    driver.get(f"{base_url}/lettings/{letting_id}")
    wait_for_page_load(driver, timeout=30)
    wait_for_elements_stable(driver, "div.p-4", timeout=30)
    soup = BeautifulSoup(driver.page_source, _HTML_PARSER)

    proposals = []
    for container in soup.find_all('div', class_='p-4'):
        try:
            link = container.find('a', {'data-cy': 'cont-id'})
            if not link:
                continue

            proposal_id = link.text.strip()
            detail_url  = link.get('href', '')
            if detail_url.startswith('/'):
                detail_url = f"https://ui.bidx.com{detail_url}"

            district = ""
            grow_ps = container.find_all('p', class_='grow')
            if grow_ps:
                district = grow_ps[0].text.strip()

            description = ""
            desc_el = container.find('p', class_='text-xs',
                                     attrs={'class': lambda x: x and 'line-clamp' in x})
            if not desc_el:
                for p in container.find_all('p', class_='text-midnight-400'):
                    pt = p.text.strip()
                    if pt and 'Project ID' not in pt:
                        desc_el = p
                        break
            if desc_el:
                description = desc_el.text.strip()

            if proposal_id and detail_url:
                proposals.append({
                    'proposal_id': proposal_id,
                    'district':    district,
                    'description': description,
                    'detail_url':  detail_url,
                })
        except:
            continue
    return proposals


def scan_proposal_items(driver, proposal_detail_url, proposal_id, district,
                        project_description, letting_date,
                        milling_keywords, grinding_items, grooving_items):
    """
    Visit one proposal page and return (milling_matches, grinding_matches).

    Milling  : keyword search against item number + description text (OR logic).
    Grinding : item-number substring match against configured item codes.
    Both lists share the same base record structure; grinding adds a 'type' key.
    """
    driver.get(proposal_detail_url)
    wait_for_page_load(driver, timeout=30)
    wait_for_elements_stable(driver, "cw-bidx-proposal-item", timeout=30)
    soup = BeautifulSoup(driver.page_source, _HTML_PARSER)

    milling_matches  = []
    grinding_matches = []

    for item_el in soup.find_all('cw-bidx-proposal-item'):
        try:
            # Item number
            item_number = ""
            num_div = item_el.find('div', class_='text-nowrap')
            if num_div:
                item_number = num_div.get_text(strip=True)

            # Description — check class list explicitly (avoids false positives from nested lambdas)
            description_text = ""
            for div in item_el.find_all('div', class_='font-medium'):
                cls = div.get('class', [])
                if ('min-h-6' not in cls and 'max-w-60' not in cls
                        and 'grow' in cls and 'basis-44' in cls and 'items-center' in cls):
                    description_text = div.get_text(strip=True)
                    break

            # Quantity / unit from structured container
            quantity_value = unit_value = ""
            qty_parent = item_el.find('div', attrs={
                'class': lambda x: x and 'min-w-32' in x and 'basis-32' in x
            })
            if qty_parent:
                qty_container = qty_parent.find('div', attrs={
                    'class': lambda x: x and 'w-full' in x and 'items-center' in x
                })
                if qty_container:
                    child_divs = qty_container.find_all('div', recursive=False)
                    if len(child_divs) >= 2:
                        quantity_value = child_divs[0].get_text(strip=True)
                        unit_value     = child_divs[1].get_text(strip=True)
                    elif len(child_divs) == 1:
                        quantity_value = child_divs[0].get_text(strip=True)

            # Regex fallback when qty/unit is embedded in the description string
            cleaned_description = description_text
            if not quantity_value or not unit_value:
                m = re.search(r'([\d,]+)\s*([A-Z]+)\s*$', cleaned_description.strip())
                if m:
                    quantity_value      = m.group(1)
                    unit_value          = m.group(2)
                    cleaned_description = cleaned_description[:m.start()].strip()

            base_record = {
                'letting_date':        letting_date,
                'proposal_id':         proposal_id,
                'district':            district,
                'project_description': project_description,
                'item_number':         item_number,
                'description':         cleaned_description,
                'unit':                unit_value,
                'quantity':            quantity_value,
            }

            # Milling: match keyword against item number OR description
            if milling_keywords:
                search_text = f"{item_number} {description_text}".upper()
                if any(kw in search_text for kw in milling_keywords):
                    milling_matches.append(base_record)

            # Grinding/Grooving: match configured item codes against item number
            if grinding_items or grooving_items:
                match_type = None
                for gi in grinding_items:
                    if gi.upper() in item_number.upper():
                        match_type = "Grinding"
                        break
                if not match_type:
                    for gvi in grooving_items:
                        if gvi.upper() in item_number.upper():
                            match_type = "Grooving"
                            break
                if match_type:
                    grinding_matches.append({**base_record, 'type': match_type})

        except:
            continue

    return milling_matches, grinding_matches


# ── EXCEL WRITING ──────────────────────────────────────────────────────────────

def _prepare_dataframe(results, col_order, headers):
    """Convert a list of result dicts into a formatted DataFrame."""
    if not results:
        return pd.DataFrame(columns=headers)
    df = pd.DataFrame(results)
    df['project_description'] = df['project_description'].apply(
        lambda x: (x[:50] + "...") if isinstance(x, str) and len(x) > 50 else x
    )
    df = df[col_order]
    df.columns = headers
    return df


def write_sheet(results, output_file, sheet_name, col_order, headers):
    """Write results to a named sheet in an output Excel workbook."""
    df   = _prepare_dataframe(results, col_order, headers)
    mode = 'a' if os.path.exists(output_file) else 'w'
    kw   = {'engine': 'openpyxl', 'mode': mode}
    if mode == 'a':
        kw['if_sheet_exists'] = 'replace'
    with pd.ExcelWriter(output_file, **kw) as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)


def autofit_all_sheets(filename, sheet_names):
    """Auto-fit column widths for all named sheets using openpyxl."""
    try:
        wb = load_workbook(filename)
        for sn in sheet_names:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            for col in ws.columns:
                max_length = max(
                    (len(str(cell.value)) for cell in col if cell.value), default=0
                )
                ws.column_dimensions[col[0].column_letter].width = (
                    min((max_length + 2) * 1.2, 60) if max_length else 10
                )
        wb.save(filename)
    except Exception as e:
        print(f"Warning: Manual autofit failed: {e}")


def merge_and_center_proposal_groups(filename, sheet_name, proposal_id_col, n_merge_cols):
    """
    Merge repeated header columns (Letting Date → Project Description) for
    rows that share the same Proposal ID, and center-align merged cells.

    proposal_id_col : 1-based column index of Proposal ID
    n_merge_cols    : how many leading columns to merge (e.g. 4 for milling, 5 for grinding)
    """
    wb = load_workbook(filename)
    ws = wb[sheet_name]
    merge_cols = list(range(1, n_merge_cols + 1))
    start = 2
    while start <= ws.max_row:
        pid = ws.cell(row=start, column=proposal_id_col).value
        end = start
        while (end + 1 <= ws.max_row
               and ws.cell(row=end + 1, column=proposal_id_col).value == pid):
            end += 1
        if end > start:
            for col in merge_cols:
                ws.merge_cells(start_row=start, start_column=col,
                               end_row=end,   end_column=col)
                ws.cell(row=start, column=col).alignment = Alignment(
                    horizontal='left', vertical='center'
                )
        start = end + 1
    wb.save(filename)


# ── ARCHIVE FUNCTIONS ──────────────────────────────────────────────────────────

def load_archived_proposal_ids(archive_file):
    """Return the set of Proposal IDs already saved in an archive file."""
    if not os.path.exists(archive_file):
        print("  No archive file found — first run.")
        return set()
    try:
        df = pd.read_excel(archive_file)
        if 'Proposal ID' in df.columns:
            ids = set(df['Proposal ID'].dropna().unique())
            print(f"  OK - Loaded {len(ids)} archived proposal IDs")
            return ids
        print("  Archive has no 'Proposal ID' column — starting fresh.")
        return set()
    except Exception as e:
        print(f"  Warning: Could not read archive: {e}")
        return set()


def filter_new_proposals(all_results, archived_ids):
    """Split results into (new, already_known) based on archived proposal IDs."""
    if not archived_ids:
        return list(all_results), []
    new, known = [], []
    for r in all_results:
        (known if r['proposal_id'] in archived_ids else new).append(r)
    return new, known


def append_to_archive(new_results, archive_file, col_order, headers):
    """Append new results to an archive Excel file, creating it if it doesn't exist."""
    if not new_results:
        return
    df_new = _prepare_dataframe(new_results, col_order, headers)
    if os.path.exists(archive_file):
        try:
            df_existing = pd.read_excel(archive_file)
            pd.concat([df_existing, df_new], ignore_index=True).to_excel(archive_file, index=False)
            print(f"  OK - Appended {len(new_results)} rows to {archive_file}")
        except Exception as e:
            print(f"  Warning: Could not append ({e}) — overwriting archive.")
            df_new.to_excel(archive_file, index=False)
            print(f"  OK - Created {archive_file} with {len(new_results)} rows")
    else:
        df_new.to_excel(archive_file, index=False)
        print(f"  OK - Created {archive_file} with {len(new_results)} rows")


# ── MAIN ───────────────────────────────────────────────────────────────────────

def main():
    print("\n" + "=" * 80)
    print("BIDX FULL SCRAPER  —  MILLING  +  GRINDING / GROOVING")
    print("=" * 80 + "\n")

    overall_start = time.time()

    username, password = load_credentials()
    print(f"Credentials loaded for: {username}")

    milling_dots  = load_milling_config()
    grinding_dots = load_grinding_config()
    dots = build_dot_list(milling_dots, grinding_dots)
    print(f"DOTs to scrape: {len(dots)}  "
          f"({len(milling_dots)} milling config, {len(grinding_dots)} grinding config)\n")

    # Accumulators (keyed by sheet name = DOT name truncated to 31 chars)
    all_results     = {}   # sheet -> {'milling': [...], 'grinding': [...]}
    archive_updates = {}   # archive_file -> (results, col_order, headers)

    driver = initialize_browser(username, password)

    for dot_idx, dot in enumerate(dots, 1):
        # Periodic browser restart to avoid memory / session issues
        if dot_idx > 1 and (dot_idx - 1) % RESTART_EVERY_N_DOTS == 0:
            print("\n" + "!" * 80)
            print(f"RESTARTING BROWSER (every {RESTART_EVERY_N_DOTS} DOTs)...")
            print("!" * 80 + "\n")
            driver.quit()
            time.sleep(2)
            driver = initialize_browser(username, password, is_restart=True)

        dot_name       = dot['name']
        url_slug       = dot['url_slug']
        safe_name      = dot_name.replace(' ', '_').replace('/', '_')
        sheet          = dot_name[:31]
        milling_kws    = dot['milling_keywords']
        grinding_items = dot['grinding_items']
        grooving_items = dot['grooving_items']

        has_milling  = bool(milling_kws)
        has_grinding = bool(grinding_items or grooving_items)

        print("\n" + "=" * 80)
        print(f"DOT {dot_idx}/{len(dots)}: {dot_name}")
        modes = ([("Milling")    if has_milling  else None] +
                 [("Grinding/Grooving") if has_grinding else None])
        print(f"Modes: {', '.join(m for m in modes if m) or 'NONE'}")
        print("=" * 80)

        if not has_milling and not has_grinding:
            print("Skipping — no keywords or item numbers configured.\n")
            continue

        dot_start        = time.time()
        milling_archive  = f"bidx_milling_{safe_name}_archive.xlsx"
        grinding_archive = f"bidx_grinding_{safe_name}_archive.xlsx"

        try:
            lettings_url = f"https://ui.bidx.com/{url_slug}/lettings"
            print(f"Navigating to {lettings_url}")
            driver.get(lettings_url)
            wait_for_page_load(driver, timeout=30)
            base_url = f"https://ui.bidx.com/{url_slug}"
            print("OK - Page loaded\n")

            print("Loading archives...")
            milling_archived  = (load_archived_proposal_ids(milling_archive)
                                 if has_milling  else set())
            grinding_archived = (load_archived_proposal_ids(grinding_archive)
                                 if has_grinding else set())
            print()

            lettings = extract_lettings_from_page(driver)
            if not lettings:
                print("WARNING: No lettings found. Skipping.\n")
                continue
        except Exception as e:
            print(f"ERROR processing {dot_name}: {e}\nSkipping.\n")
            continue

        print(f"Processing {len(lettings)} lettings...\n")
        dot_milling  = []
        dot_grinding = []

        for letting_idx, letting in enumerate(lettings):
            letting_id   = letting['letting_id']
            letting_date = letting['letting_date']

            print(f"  [{letting_idx + 1}/{len(lettings)}] {letting_id} ({letting_date})... ",
                  end='', flush=True)

            proposals = navigate_to_letting_and_extract_proposals(driver, base_url, letting_id)
            if not proposals:
                print("no proposals")
                continue

            m_hits = g_hits = 0
            for proposal in proposals:
                ml, gr = scan_proposal_items(
                    driver,
                    proposal['detail_url'],
                    proposal['proposal_id'],
                    proposal['district'],
                    proposal['description'],
                    letting_date,
                    milling_kws    if has_milling  else [],
                    grinding_items if has_grinding else [],
                    grooving_items if has_grinding else [],
                )
                m_hits += len(ml)
                g_hits += len(gr)
                dot_milling.extend(ml)
                dot_grinding.extend(gr)

            print(f"{len(proposals)} proposals | milling={m_hits}  grinding={g_hits}")

        # Filter to new-only
        new_milling,  known_milling  = filter_new_proposals(dot_milling,  milling_archived)
        new_grinding, known_grinding = filter_new_proposals(dot_grinding, grinding_archived)

        print(f"\n  Milling  : {len(dot_milling)} total | "
              f"{len(new_milling)} new | {len(known_milling)} archived")
        print(f"  Grinding : {len(dot_grinding)} total | "
              f"{len(new_grinding)} new | {len(known_grinding)} archived\n")

        if new_milling or new_grinding:
            all_results[sheet] = {'milling': new_milling, 'grinding': new_grinding}

        if new_milling:
            archive_updates[milling_archive]  = (new_milling,  _MILLING_COLS,  _MILLING_HDRS)
        if new_grinding:
            archive_updates[grinding_archive] = (new_grinding, _GRINDING_COLS, _GRINDING_HDRS)

        dot_elapsed = time.time() - dot_start
        print(f"  DOT done in {dot_elapsed:.1f}s")
        print("-" * 80)

    driver.quit()

    # ── Write Excel outputs ──────────────────────────────────────────────────
    if all_results:
        print("\n" + "=" * 80)
        print("WRITING EXCEL OUTPUT")
        print("=" * 80)

        milling_sheets  = [s for s, v in all_results.items() if v['milling']]
        grinding_sheets = [s for s, v in all_results.items() if v['grinding']]

        for sheet, v in all_results.items():
            if v['milling']:
                print(f"  Writing milling  sheet '{sheet}'...")
                write_sheet(v['milling'],  MILLING_OUTPUT,  sheet,
                            _MILLING_COLS,  _MILLING_HDRS)
            if v['grinding']:
                print(f"  Writing grinding sheet '{sheet}'...")
                write_sheet(v['grinding'], GRINDING_OUTPUT, sheet,
                            _GRINDING_COLS, _GRINDING_HDRS)

        print("Applying cell merging...")
        for sheet, v in all_results.items():
            if v['milling']:
                # Milling: Proposal ID in col 2, merge cols 1-4
                merge_and_center_proposal_groups(
                    MILLING_OUTPUT, sheet, proposal_id_col=2, n_merge_cols=4)
            if v['grinding']:
                # Grinding: Proposal ID in col 3 (Type is col 1), merge cols 1-5
                merge_and_center_proposal_groups(
                    GRINDING_OUTPUT, sheet, proposal_id_col=3, n_merge_cols=5)

        print("Autofitting columns...")
        if milling_sheets:
            autofit_all_sheets(MILLING_OUTPUT, milling_sheets)
        if grinding_sheets:
            autofit_all_sheets(GRINDING_OUTPUT, grinding_sheets)

        print(f"\nOK - {MILLING_OUTPUT}")
        print(f"OK - {GRINDING_OUTPUT}\n")

    # ── Update archives ──────────────────────────────────────────────────────
    if archive_updates:
        print("=" * 80)
        print("UPDATING ARCHIVES")
        print("=" * 80)
        for archive_file, (results, col_order, headers) in archive_updates.items():
            append_to_archive(results, archive_file, col_order, headers)
        print()

    # ── Final summary ────────────────────────────────────────────────────────
    overall_elapsed = time.time() - overall_start
    total_milling   = sum(len(v['milling'])  for v in all_results.values())
    total_grinding  = sum(len(v['grinding']) for v in all_results.values())

    print("=" * 80)
    print("BIDX FULL SCRAPER — SUMMARY")
    print("=" * 80)
    print(f"Total time    : {overall_elapsed:.1f}s")
    print(f"DOTs scraped  : {len(all_results)}")
    print(f"Milling rows  : {total_milling}  → {MILLING_OUTPUT}")
    print(f"Grinding rows : {total_grinding} → {GRINDING_OUTPUT}")
    print("=" * 80 + "\n")

    # ── Azure Blob Storage (set BLOB_ENABLED=true; Power Automate copies to SharePoint) ─
    try:
        from blob_publish import blob_enabled, publish_to_blob
        if blob_enabled():
            publish_to_blob(MILLING_OUTPUT, GRINDING_OUTPUT)
        else:
            print("Blob publish skipped (BLOB_ENABLED is not true).")
    except Exception as e:
        print(f"WARNING: Blob publish failed: {e}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\nInterrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
